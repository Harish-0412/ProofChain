"""Deterministic evidence safety inspection and non-destructive quarantine."""

from __future__ import annotations

import hashlib
import re
import shutil
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from proofchain.core.paths import get_run_dir
from proofchain.schemas.production import EvidenceSecurityFinding, SecurityInput


PROMPT_PATTERNS = (
    "ignore previous instructions",
    "system prompt",
    "override policy",
    "execute tool",
    "developer message",
)
PII_PATTERNS = (
    re.compile(r"\b\d{3}-\d{2}-\d{4}\b"),
    re.compile(r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", re.IGNORECASE),
)
DANGEROUS_SUFFIXES = {".exe", ".dll", ".bat", ".cmd", ".ps1", ".js", ".vbs", ".scr"}


def inspect_evidence(request: SecurityInput) -> list[EvidenceSecurityFinding]:
    allowed_roots = [Path(root).resolve() for root in request.allowed_roots]
    results: list[EvidenceSecurityFinding] = []
    for raw_path in request.evidence_paths:
        path = Path(raw_path).resolve()
        findings: list[str] = []
        restrictions: list[str] = []
        decision = "ALLOW"
        digest: str | None = None
        quarantine_reference: str | None = None

        if allowed_roots and not any(path == root or root in path.parents for root in allowed_roots):
            findings.append("path_outside_allowed_roots")
            decision = "REJECT"
        elif not path.is_file():
            findings.append("file_missing")
            decision = "REJECT"
        else:
            size = path.stat().st_size
            digest = _file_hash(path)
            if size > request.max_file_bytes:
                findings.append("oversized_file")
                decision = "QUARANTINE"
            if path.suffix.lower() in DANGEROUS_SUFFIXES:
                findings.append("dangerous_executable_type")
                decision = "QUARANTINE"
            sample = path.read_bytes()[:2_000_000]
            lowered = sample.lower()
            if b"eicar-standard-antivirus-test-file" in lowered:
                findings.append("malware_signature_detected")
                decision = "QUARANTINE"
            text = sample.decode("utf-8", errors="ignore")
            if any(pattern in text.lower() for pattern in PROMPT_PATTERNS):
                findings.append("prompt_injection_content")
                restrictions.append("content_must_not_influence_agent_instructions")
                if decision == "ALLOW":
                    decision = "ALLOW_WITH_RESTRICTIONS"
            if any(pattern.search(text) for pattern in PII_PATTERNS):
                findings.append("possible_pii")
                restrictions.append("redacted_derivative_required_for_external_use")
                if decision in {"ALLOW", "ALLOW_WITH_RESTRICTIONS"}:
                    decision = "REDACT_DERIVATIVE_REQUIRED"
            if zipfile.is_zipfile(path):
                archive_findings = _inspect_archive(path)
                findings.extend(archive_findings)
                if archive_findings:
                    decision = "QUARANTINE"
            if path.suffix.lower() in {".xlsx", ".xlsm"}:
                spreadsheet_findings = _inspect_spreadsheet(path)
                findings.extend(spreadsheet_findings)
                if spreadsheet_findings and decision == "ALLOW":
                    decision = "ALLOW_WITH_RESTRICTIONS"

        if decision == "QUARANTINE" and request.quarantine_enabled and path.is_file():
            quarantine_dir = get_run_dir(request.workflow.run_id) / "quarantine"
            quarantine_dir.mkdir(parents=True, exist_ok=True)
            destination = quarantine_dir / f"{digest or 'unknown'}{path.suffix.lower()}"
            if not destination.exists():
                shutil.copy2(path, destination)
            quarantine_reference = str(destination.resolve())
        results.append(
            EvidenceSecurityFinding(
                path=str(path),
                decision=decision,
                findings=sorted(set(findings)),
                restrictions=sorted(set(restrictions)),
                sha256=digest,
                quarantine_reference=quarantine_reference,
            )
        )
    return results


def _inspect_archive(path: Path) -> list[str]:
    findings: list[str] = []
    try:
        with zipfile.ZipFile(path) as archive:
            total = 0
            for info in archive.infolist():
                member = Path(info.filename)
                if member.is_absolute() or ".." in member.parts:
                    findings.append("archive_path_traversal")
                total += info.file_size
                if info.compress_size and info.file_size / info.compress_size > 100:
                    findings.append("archive_compression_bomb_risk")
                if total > 500 * 1024 * 1024:
                    findings.append("archive_expansion_limit_exceeded")
    except (OSError, zipfile.BadZipFile):
        findings.append("invalid_archive")
    return findings


def _inspect_spreadsheet(path: Path) -> list[str]:
    findings: list[str] = []
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        if any(sheet.sheet_state != "visible" for sheet in workbook.worksheets):
            findings.append("hidden_worksheet")
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
                        findings.append("spreadsheet_formula_content")
                        return sorted(set(findings))
    except Exception:
        findings.append("spreadsheet_inspection_failed")
    return sorted(set(findings))


def _file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()

