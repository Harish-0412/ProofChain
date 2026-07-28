"""XLSX and CSV extraction with source coordinates."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from proofchain.core.enums import ExtractionStatus
from proofchain.schemas.classification import ExtractionResult
from proofchain.schemas.common import SourceReference


class SpreadsheetExtractor:
    def extract_xlsx(self, path: Path, evidence_id: str) -> ExtractionResult:
        workbook = load_workbook(path, data_only=True, read_only=True)
        tables: list[dict[str, Any]] = []
        references: list[SourceReference] = []
        sheet_names = list(workbook.sheetnames)

        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            row_count = max(1, len(rows))
            column_count = max(
                1,
                max((len(row) for row in rows), default=1),
            )
            tables.append({"sheet_name": sheet.title, "rows": rows})
            references.append(
                SourceReference(
                    document=evidence_id,
                    sheet_name=sheet.title,
                    cell_range=f"A1:{get_column_letter(column_count)}{row_count}",
                )
            )
        workbook.close()
        return ExtractionResult(
            extraction_status=ExtractionStatus.SUCCESS,
            extractor_used="openpyxl",
            tables=tables,
            sheet_names=sheet_names,
            page_references=references,
            extraction_confidence=0.99,
        )

    def extract_csv(self, path: Path, evidence_id: str) -> ExtractionResult:
        return self.extract_delimited(path, evidence_id, delimiter=",", sheet_name="CSV")

    def extract_tsv(self, path: Path, evidence_id: str) -> ExtractionResult:
        return self.extract_delimited(path, evidence_id, delimiter="\t", sheet_name="TSV")

    @staticmethod
    def extract_delimited(
        path: Path,
        evidence_id: str,
        *,
        delimiter: str,
        sheet_name: str,
    ) -> ExtractionResult:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle, delimiter=delimiter))
        return ExtractionResult(
            extraction_status=ExtractionStatus.SUCCESS,
            extractor_used="csv" if delimiter == "," else "tsv",
            tables=[{"sheet_name": sheet_name, "rows": rows}],
            sheet_names=[sheet_name],
            page_references=[
                SourceReference(
                    document=evidence_id,
                    sheet_name=sheet_name,
                    cell_range=f"A1:{len(rows)}",
                )
            ],
            extraction_confidence=0.98,
        )
