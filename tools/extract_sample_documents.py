from __future__ import annotations

import json
import re
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SAMPLE_ROOT = ROOT / "sample_data"
OUTPUT_PATH = SAMPLE_ROOT / "extraction_outputs" / "extracted_fields.json"


FIELD_MAP = {
    "Event ID": "event_id",
    "Accreditation Requirement": "requirement",
    "Mapped Requirement": "mapped_requirement",
    "Department": "department",
    "Academic Year": "academic_year",
    "Event Title": "event_title",
    "Event Date": "event_date",
    "Coordinator": "coordinator",
    "Reported Participant Count": "reported_participant_count",
    "Signature Present": "signature_present",
    "Approval Status": "approval_status",
}


def extract_pdf(path: Path) -> dict:
    with pdfplumber.open(str(path)) as pdf:
        pages = [page.extract_text() or "" for page in pdf.pages]
    text = "\n".join(pages)
    fields = {}
    references = {}
    for label, field in FIELD_MAP.items():
        pattern = re.compile(rf"{re.escape(label)}\s+(.+)")
        match = pattern.search(text)
        if match:
            value = match.group(1).strip()
            if field == "reported_participant_count":
                value = int(re.search(r"\d+", value).group(0))
            fields[field] = value
            references[field] = "page 1"

    document_type = "event_report"
    if "Approval Document" in text:
        document_type = "approval_document"
    if "Certificate of Participation" in text:
        document_type = "certificate"

    return {
        "file": str(path.relative_to(ROOT)),
        "document_type": document_type,
        "extracted_fields": fields,
        "source_references": references,
        "confidence": 0.9 if fields else 0.45,
    }


def extract_xlsx(path: Path) -> dict:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    headers = [sheet.cell(row=3, column=col).value for col in range(1, sheet.max_column + 1)]
    records = []
    for row in range(4, sheet.max_row + 1):
        record = {
            headers[col - 1]: sheet.cell(row=row, column=col).value
            for col in range(1, sheet.max_column + 1)
            if headers[col - 1]
        }
        if record.get("Roll Number"):
            records.append(record)

    roll_numbers = [record["Roll Number"] for record in records]
    unique_rolls = sorted(set(roll_numbers))
    duplicates = sorted({roll for roll in roll_numbers if roll_numbers.count(roll) > 1})
    first = records[0] if records else {}

    return {
        "file": str(path.relative_to(ROOT)),
        "document_type": "attendance_sheet",
        "extracted_fields": {
            "event_id": first.get("Event ID"),
            "event_title": first.get("Event Title"),
            "department": first.get("Department"),
            "attendance_rows": len(records),
            "unique_student_count": len(unique_rolls),
            "duplicate_roll_numbers": duplicates,
        },
        "source_references": {
            "unique_student_count": "Attendance!B4:B" + str(sheet.max_row),
            "duplicate_roll_numbers": "Attendance!B4:B" + str(sheet.max_row),
        },
        "confidence": 0.97,
    }


def main() -> None:
    results = []
    for path in SAMPLE_ROOT.glob("departments/**/*.pdf"):
        results.append(extract_pdf(path))
    for path in SAMPLE_ROOT.glob("departments/**/*.xlsx"):
        results.append(extract_xlsx(path))
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(results, indent=2), encoding="utf-8")
    print(f"Extracted {len(results)} documents into {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
