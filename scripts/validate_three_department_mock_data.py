"""Validate structure, identities, checksums, attendance, and format coverage."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EXPECTED_DEPARTMENTS = {"AIML", "AIDS", "CSE"}
EXPECTED_NATIVE_FORMATS = {
    ".pdf",
    ".xlsx",
    ".csv",
    ".tsv",
    ".docx",
    ".txt",
    ".md",
    ".json",
    ".xml",
    ".html",
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def attendance_records(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.worksheets[0]
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
    else:
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle, delimiter=delimiter))
    if not rows:
        return []
    headers = [str(value) for value in rows[0]]
    return [
        {headers[index]: value for index, value in enumerate(row)}
        for row in rows[1:]
        if any(value not in (None, "") for value in row)
    ]


def validate(root: Path) -> dict[str, Any]:
    errors: list[str] = []
    checks: list[dict[str, Any]] = []
    manifest = json.loads((root / "dataset_manifest.json").read_text(encoding="utf-8"))
    payload = json.loads((root / "workbook_payload.json").read_text(encoding="utf-8"))

    departments = set(manifest["departments"])
    if departments != EXPECTED_DEPARTMENTS:
        errors.append(f"Departments are {sorted(departments)}, expected {sorted(EXPECTED_DEPARTMENTS)}.")
    checks.append({"name": "department_scope", "passed": departments == EXPECTED_DEPARTMENTS})

    students = payload["students"]
    roll_numbers = [student["roll_number"] for student in students]
    emails = [student["email"] for student in students]
    student_ok = len(students) == 90 and len(set(roll_numbers)) == 90 and len(set(emails)) == 90
    if not student_ok:
        errors.append("Student identities are not exactly 90 unique roll numbers and emails.")
    checks.append({"name": "student_identity_uniqueness", "passed": student_ok})

    for department in sorted(EXPECTED_DEPARTMENTS):
        department_students = [
            student for student in students if student["department"] == department
        ]
        female_count = sum(student["gender"] == "Female" for student in department_students)
        male_count = sum(student["gender"] == "Male" for student in department_students)
        passed = len(department_students) == 30 and female_count == 15 and male_count == 15
        if not passed:
            errors.append(
                f"{department} has {len(department_students)} students, "
                f"{female_count} female and {male_count} male."
            )
        checks.append(
            {
                "name": f"{department.lower()}_student_distribution",
                "passed": passed,
                "student_count": len(department_students),
                "female_count": female_count,
                "male_count": male_count,
            }
        )

    checksum_failures = []
    for record in manifest["files"]:
        path = root / record["relative_path"]
        if not path.exists() or sha256(path) != record["sha256"]:
            checksum_failures.append(record["relative_path"])
    if checksum_failures:
        errors.append(f"Checksum failures: {checksum_failures}")
    checks.append(
        {
            "name": "manifest_checksums",
            "passed": not checksum_failures,
            "checked_files": len(manifest["files"]),
        }
    )

    events = manifest["events"]
    event_ids = [event["event_id"] for event in events]
    event_ok = len(events) == 15 and len(set(event_ids)) == 15
    if not event_ok:
        errors.append(f"Expected 15 unique events, found {len(events)} records.")
    checks.append({"name": "event_identity", "passed": event_ok})

    attendance_failures = []
    for event in events:
        attendance_path = root / event["files"]["attendance_sheet"]
        records = attendance_records(attendance_path)
        unique_rolls = {
            str(record.get("Roll Number", "")).strip()
            for record in records
            if record.get("Roll Number")
        }
        departments_found = {
            str(record.get("Department", "")).strip() for record in records
        }
        events_found = {str(record.get("Event ID", "")).strip() for record in records}
        if (
            len(records) != 30
            or len(unique_rolls) != 30
            or departments_found != {event["department"]}
            or events_found != {event["event_id"]}
        ):
            attendance_failures.append(event["event_id"])
    if attendance_failures:
        errors.append(f"Attendance reconciliation failed: {attendance_failures}")
    checks.append(
        {
            "name": "attendance_reconciliation",
            "passed": not attendance_failures,
            "checked_events": len(events),
        }
    )

    evidence_files = [
        path for path in (root / "departments").rglob("*") if path.is_file()
    ]
    evidence_count_ok = len(evidence_files) == 75
    if not evidence_count_ok:
        errors.append(f"Expected 75 evidence files, found {len(evidence_files)}.")
    checks.append(
        {
            "name": "evidence_file_count",
            "passed": evidence_count_ok,
            "count": len(evidence_files),
        }
    )

    folder_counts: dict[str, int] = {}
    for path in evidence_files:
        evidence_type = path.parent.name
        folder_counts[evidence_type] = folder_counts.get(evidence_type, 0) + 1
    expected_folder_counts = {
        "event_reports": 15,
        "attendance_sheets": 15,
        "approval_documents": 15,
        "certificates": 15,
        "photos": 15,
    }
    folder_ok = folder_counts == expected_folder_counts
    if not folder_ok:
        errors.append(f"Evidence folder counts are {folder_counts}.")
    checks.append(
        {
            "name": "evidence_bundle_shape",
            "passed": folder_ok,
            "counts": folder_counts,
        }
    )

    represented = {path.suffix.lower() for path in evidence_files}
    native_coverage_ok = EXPECTED_NATIVE_FORMATS.issubset(represented)
    if not native_coverage_ok:
        errors.append(
            f"Missing native formats: {sorted(EXPECTED_NATIVE_FORMATS - represented)}"
        )
    checks.append(
        {
            "name": "native_format_coverage",
            "passed": native_coverage_ok,
            "formats": sorted(represented),
        }
    )

    return {
        "dataset_id": manifest["dataset_id"],
        "valid": not errors,
        "summary": {
            "departments": len(departments),
            "students": len(students),
            "events": len(events),
            "evidence_files": len(evidence_files),
            "checks_passed": sum(check["passed"] for check in checks),
            "checks_total": len(checks),
        },
        "checks": checks,
        "errors": errors,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--root",
        type=Path,
        default=Path("sample_data/mock_institution"),
    )
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    report = validate(args.root)
    output = args.output or args.root / "validation_report.json"
    output.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, indent=2))
    raise SystemExit(0 if report["valid"] else 1)


if __name__ == "__main__":
    main()
